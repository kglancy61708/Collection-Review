"""
Excel workbook writer using openpyxl.
Produces a two-sheet workbook: main collections review + summary.
"""
from __future__ import annotations

import io
import datetime
from typing import Optional

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from .logic import (
    AccountSummary, BUCKETS, AGED_BUCKETS, compute_bucket_ranges,
    STATUS_SEVERITY, _is_greystar
)

# ---------------------------------------------------------------------------
# Color palette
# ---------------------------------------------------------------------------

STATUS_COLORS = {
    "Collection Letter":      ("C00000", "FFFFFF"),
    "Predemand Letter":       ("E26B0A", "FFFFFF"),
    "Cancel NP":              ("7B0000", "FFFFFF"),
    "To be Determined":       ("FFFF00", "000000"),
    "Call Customer":          ("F4B942", "FFFFFF"),
    "Skipped Invoice":        ("7030A0", "FFFFFF"),
    "Friendly Reminder Email": ("375623", "FFFFFF"),
    "Special Circumstances":  ("2E75B6", "FFFFFF"),
}

REP_COLORS = {
    "Rosas, Yoniva":   ("E2EFDA", "000000"),
    "Quilantan, Maria": ("FFF2CC", "000000"),
    "Wharton, Nancy":  ("D9E1F2", "000000"),
}

FUTURE_RESTRICTION_COLORS = {
    "Warning":     ("FF9900", "FFFFFF"),
    "Short Leash": ("FF0000", "FFFFFF"),
    "Sales":       ("70AD47", "FFFFFF"),
    "Withheld":    ("808080", "FFFFFF"),
}

ESCALATION_COLORS = {
    "Short Leash": ("FF0000", "FFFFFF"),
    "Long Leash":  ("7030A0", "FFFFFF"),
}

FINANCE_CHARGE_HEADER_COLOR = ("4472C4", "FFFFFF")
FINANCE_CHARGE_CELL_COLOR = "E8F0FE"

PREPAID_RESTRICTION_BG = "BDD7EE"  # light blue — prepaid 12/31/2026 + restriction warning

GREYSTAR_BG = "D9D9D9"
AUTOPAY_BG = "F4B942"
CURRENT_ONLY_BG = "F2F2F2"
CURRENT_ONLY_TEXT = "595959"
CREDIT_ADJ_BG = "E2EFDA"
CREDIT_ADJ_TEXT = "375623"

HEADER_BG = "2F5496"
HEADER_TEXT = "FFFFFF"

SECTION_SEPARATOR_BG = "D9D9D9"

# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

OUTPUT_COLUMNS = [
    "Collect As",
    "Business Unit",
    "Category",
    "Current Collections Status",
    "Collection Escalation Status",
    "Account Restricted",
    "Future Restriction",
    "Fortis Autopay Enrollment",
    "Current Balance",
    "30+ Balance",
    "60+ Balance",
    "90+ Balance",
    "120+ Balance",
    "150+ Balance",
    "180+ Balance",
    "Finance Charges",
    "Total Open Balance",
    "Total Open Credits",
    "Credit Adjustment",
    "# Invoices",
    "Oldest Bucket w/ Balance",
    "Suggested Status",  # header will include month name
    "Collections Assignment",
]

MONEY_COLUMNS = {
    "Current Balance", "30+ Balance", "60+ Balance", "90+ Balance",
    "120+ Balance", "150+ Balance", "180+ Balance", "Finance Charges",
    "Total Open Balance", "Total Open Credits",
}


def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _make_font(bold=False, italic=False, color="000000") -> Font:
    return Font(bold=bold, italic=italic, color=color)


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def build_workbook(
    accounts: list[AccountSummary],
    report_month: int,
    report_year: int,
) -> bytes:
    """Build the Excel workbook and return as bytes."""
    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "Collections Review"
    ws_summary = wb.create_sheet("Summary")

    month_name = datetime.date(report_year, report_month, 1).strftime("%B %Y")

    _write_main_sheet(ws_main, accounts, report_month, report_year, month_name)
    _write_summary_sheet(ws_summary, accounts, report_month, report_year, month_name)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Main sheet
# ---------------------------------------------------------------------------

def _write_main_sheet(ws, accounts, report_month, report_year, month_name):
    # Build column headers (replace "Suggested Status" with month-specific)
    headers = [c if c != "Suggested Status" else f"Suggested {month_name} Status"
               for c in OUTPUT_COLUMNS]

    col_idx = {h: i + 1 for i, h in enumerate(headers)}
    status_col_name = f"Suggested {month_name} Status"

    # Write header row
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _make_font(bold=True, color=HEADER_TEXT)
        if h == "Finance Charges":
            cell.fill = _make_fill(FINANCE_CHARGE_HEADER_COLOR[0])
        else:
            cell.fill = _make_fill(HEADER_BG)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ws.freeze_panes = "A2"

    row_num = 2
    prev_section = None

    for acc in accounts:
        # Insert section separator row
        if acc.section != prev_section and prev_section is not None:
            label = "── Current Only ──" if acc.section == "current_only" else "── Autopay (Aged) ──"
            sep_cell = ws.cell(row=row_num, column=1, value=label)
            sep_cell.font = _make_font(bold=True)
            sep_cell.fill = _make_fill(SECTION_SEPARATOR_BG)
            ws.merge_cells(
                start_row=row_num, start_column=1,
                end_row=row_num, end_column=len(headers)
            )
            row_num += 1
        prev_section = acc.section

        row_data = _account_to_row(acc, status_col_name)
        _write_data_row(ws, row_num, headers, row_data, acc, status_col_name)
        row_num += 1

    # Auto-width columns
    for col in range(1, len(headers) + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)

    # Set row height for header
    ws.row_dimensions[1].height = 30


def _account_to_row(acc: AccountSummary, status_col_name: str) -> dict:
    return {
        "Collect As": acc.collect_as,
        "Business Unit": acc.business_unit,
        "Category": acc.category,
        "Current Collections Status": acc.current_collections_status,
        "Collection Escalation Status": acc.collection_escalation_status,
        "Account Restricted": acc.account_restricted,
        "Future Restriction": acc.future_restriction,
        "Fortis Autopay Enrollment": acc.fortis_autopay_enrollment,
        "Current Balance": acc.balances["Current"] or None,
        "30+ Balance": acc.balances["30+"] or None,
        "60+ Balance": acc.balances["60+"] or None,
        "90+ Balance": acc.balances["90+"] or None,
        "120+ Balance": acc.balances["120+"] or None,
        "150+ Balance": acc.balances["150+"] or None,
        "180+ Balance": acc.balances["180+"] or None,
        "Finance Charges": acc.balances["Finance Charges"] or None,
        "Total Open Balance": acc.total_open_balance or None,
        "Total Open Credits": acc.total_open_credits or None,
        "Credit Adjustment": acc.credit_adjustment,
        "# Invoices": acc.invoice_count,
        "Oldest Bucket w/ Balance": acc.oldest_bucket_with_balance,
        status_col_name: acc.suggested_status,
        "Collections Assignment": acc.collections_assignment,
    }


def _write_data_row(ws, row_num: int, headers: list, row_data: dict, acc: AccountSummary, status_col_name: str):
    status_col_name_key = status_col_name

    # Determine base row formatting
    is_greystar = _is_greystar(acc.collect_as)
    is_autopay_section = acc.section == "autopay"
    is_current_only = acc.section == "current_only"
    has_credit_adj = bool(acc.credit_adjustment)
    is_prepaid_restriction = acc.prepaid_restriction_flag

    for col_idx, h in enumerate(headers, 1):
        value = row_data.get(h)
        cell = ws.cell(row=row_num, column=col_idx, value=value)

        # Number format for money columns
        if h in MONEY_COLUMNS and value is not None:
            cell.number_format = '"$"#,##0.00'

        # Base row fill — prepaid restriction overrides other section fills
        if is_prepaid_restriction:
            cell.fill = _make_fill(PREPAID_RESTRICTION_BG)
            cell.font = _make_font(bold=True)
        elif is_greystar:
            cell.fill = _make_fill(GREYSTAR_BG)
            cell.font = _make_font(bold=True)
        elif is_autopay_section:
            cell.fill = _make_fill(AUTOPAY_BG)
            cell.font = _make_font(bold=True)
        elif is_current_only:
            cell.fill = _make_fill(CURRENT_ONLY_BG)
            cell.font = _make_font(italic=True, color=CURRENT_ONLY_TEXT)
        elif has_credit_adj:
            cell.font = _make_font(bold=True)

        # Column-specific overrides
        if h == status_col_name_key:
            if value and value in STATUS_COLORS:
                bg, fg = STATUS_COLORS[value]
                cell.fill = _make_fill(bg)
                cell.font = _make_font(bold=True, color=fg)

        elif h == "Collections Assignment":
            if value and value in REP_COLORS:
                bg, fg = REP_COLORS[value]
                cell.fill = _make_fill(bg)

        elif h == "Future Restriction":
            if value and value in FUTURE_RESTRICTION_COLORS:
                bg, fg = FUTURE_RESTRICTION_COLORS[value]
                cell.fill = _make_fill(bg)
                cell.font = _make_font(bold=True, color=fg)

        elif h == "Collection Escalation Status":
            if value and value in ESCALATION_COLORS:
                bg, fg = ESCALATION_COLORS[value]
                cell.fill = _make_fill(bg)
                cell.font = _make_font(bold=True, color=fg)

        elif h == "Finance Charges":
            if not is_greystar and not is_autopay_section and not is_current_only:
                cell.fill = _make_fill(FINANCE_CHARGE_CELL_COLOR)

        elif h == "Credit Adjustment":
            if has_credit_adj and value:
                cell.fill = _make_fill(CREDIT_ADJ_BG)
                cell.font = _make_font(bold=True, color=CREDIT_ADJ_TEXT)


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------

def _write_summary_sheet(ws, accounts, report_month, report_year, month_name):
    bucket_ranges = compute_bucket_ranges(report_year, report_month)

    row = 1

    # Title
    ws.cell(row=row, column=1, value=f"Collections Review Summary — {month_name}").font = _make_font(bold=True)
    ws.cell(row=row, column=1).fill = _make_fill(HEADER_BG)
    ws.cell(row=row, column=1).font = _make_font(bold=True, color=HEADER_TEXT)
    row += 2

    # --- Bucket date reference ---
    ws.cell(row=row, column=1, value="Bucket Date Ranges").font = _make_font(bold=True)
    row += 1

    bucket_header_labels = ["Bucket", "Start Date", "End Date"]
    for col, h in enumerate(bucket_header_labels, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = _make_font(bold=True, color=HEADER_TEXT)
        c.fill = _make_fill(HEADER_BG)
    row += 1

    date_fmt = "%m/%d/%Y"
    for b in ["Current", "30+", "60+", "90+", "120+", "150+", "180+"]:
        start, end = bucket_ranges[b]
        ws.cell(row=row, column=1, value=b)
        ws.cell(row=row, column=2, value=start.strftime(date_fmt))
        ws.cell(row=row, column=3, value=end.strftime(date_fmt))
        row += 1

    row += 1

    # --- Status × Rep matrix ---
    ws.cell(row=row, column=1, value="Status Counts by Rep").font = _make_font(bold=True)
    row += 1

    all_statuses = list(STATUS_SEVERITY.keys()) + ["No Action"]
    reps = ["Wharton, Nancy", "Rosas, Yoniva", "Quilantan, Maria", "", "Total"]

    # Header
    ws.cell(row=row, column=1, value="Status").font = _make_font(bold=True)
    for ci, rep in enumerate(reps, 2):
        c = ws.cell(row=row, column=ci, value=rep if rep else "Unassigned")
        c.font = _make_font(bold=True, color=HEADER_TEXT)
        c.fill = _make_fill(HEADER_BG)
    row += 1

    actionable_accounts = [a for a in accounts if a.section == "actionable"]
    for status in all_statuses:
        ws.cell(row=row, column=1, value=status)
        total = 0
        for ci, rep in enumerate(reps[:-1], 2):
            count = sum(
                1 for a in actionable_accounts
                if a.suggested_status == status and a.collections_assignment == rep
            )
            ws.cell(row=row, column=ci, value=count if count else None)
            total += count
        ws.cell(row=row, column=len(reps) + 1, value=total if total else None)
        row += 1

    # Totals row
    ws.cell(row=row, column=1, value="Total").font = _make_font(bold=True)
    for ci, rep in enumerate(reps[:-1], 2):
        count = sum(1 for a in actionable_accounts if a.collections_assignment == rep)
        ws.cell(row=row, column=ci, value=count if count else None).font = _make_font(bold=True)
    ws.cell(row=row, column=len(reps) + 1, value=len(actionable_accounts)).font = _make_font(bold=True)
    row += 2

    # --- Section counts ---
    current_only_count = sum(1 for a in accounts if a.section == "current_only")
    autopay_aged_count = sum(1 for a in accounts if a.section == "autopay")

    ws.cell(row=row, column=1, value="Current-Only Accounts:").font = _make_font(bold=True)
    ws.cell(row=row, column=2, value=current_only_count)
    row += 1
    ws.cell(row=row, column=1, value="Autopay w/ Aged Balance:").font = _make_font(bold=True)
    ws.cell(row=row, column=2, value=autopay_aged_count)
    row += 2

    # --- Rules reference ---
    ws.cell(row=row, column=1, value="Business Rules Reference").font = _make_font(bold=True)
    row += 1

    rules = [
        "Exclusions: Legal status, Collection Agency status, Goldspur, Aquasol Companies Employee",
        "Credits: SRDP custom form credits excluded from totals",
        "Finance Charges: Bucketed separately regardless of invoice date",
        "Future invoices: Excluded from all balance buckets",
        "Short Leash: Advances status one tier (e.g. FRE → Call Customer)",
        "Service And Repair: Advances status one tier independently",
        "Small balance (≤$25 in oldest bucket): Steps status down one tier",
        "Maintenance 90+/120+/150+/180+: Cancel NP → Wharton, Nancy",
        "Special Circumstances: Retain status → Wharton, Nancy",
        "Autopay: No collection action; aged balance shown in Autopay section",
        "Decron/Bridge/Richdale: Call Customer, Sales (green), Unassigned",
        "Dominium: Call Customer, Sales (green), Rosas or Quilantan",
        "Greystar: Normal bucket status, Withheld (grey), bold+grey row, Unassigned",
    ]
    for rule in rules:
        ws.cell(row=row, column=1, value=rule)
        row += 1

    # Auto-width
    for col in range(1, 7):
        col_letter = get_column_letter(col)
        max_len = 0
        for r in ws.iter_rows(min_col=col, max_col=col):
            for cell in r:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 50)
